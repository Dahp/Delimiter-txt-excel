import os
from tkinter import Tk, filedialog, simpledialog
from openpyxl import Workbook
import zipfile

def text_to_excel(input_directory, output_txt, output_other, delimiter):
    wb_txt = Workbook()
    wb_other = Workbook()
    
    def process_folder(folder_path):
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)
            if os.path.isfile(item_path):
                if item.endswith('.zip'):
                    process_archive(item_path, wb_txt, wb_other)
                elif item.endswith('.txt'):
                    process_txt(item_path, wb_txt, delimiter)
            elif os.path.isdir(item_path):
                process_folder(item_path)

    def process_archive(item_path, wb_txt, wb_other):
        with zipfile.ZipFile(item_path, 'r') as zip_ref:
            for file_name in zip_ref.namelist():
                if file_name.endswith('.txt'):
                    process_txt_from_zip(zip_ref, file_name, wb_txt, delimiter)
                elif file_name.endswith('.mck') or file_name.endswith('.usl'):
                    process_other_from_zip(zip_ref, file_name, wb_other, delimiter)

    def process_txt(item_path, wb_txt, delimiter):
        sheet_name = os.path.basename(item_path)
        sheet = wb_txt.create_sheet(title=sheet_name)
        with open(item_path, 'r', encoding='cp1251') as txt_file:
            lines = txt_file.readlines()
            for row_idx, line in enumerate(lines, start=1):
                values = line.strip().split(delimiter)
                for col_idx, cell_value in enumerate(values, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    def process_txt_from_zip(zip_ref, file_name, wb_txt, delimiter):
        sheet_name = os.path.splitext(os.path.basename(file_name))[0]
        sheet = wb_txt.create_sheet(title=sheet_name)
        with zip_ref.open(file_name, 'r') as txt_file:
            lines = txt_file.read().decode('cp1251').splitlines()
            for row_idx, line in enumerate(lines, start=1):
                values = line.strip().split(delimiter)
                for col_idx, cell_value in enumerate(values, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    def process_other_from_zip(zip_ref, file_name, wb_other, delimiter):
        sheet_name = os.path.splitext(os.path.basename(file_name))[0]
        sheet = wb_other.create_sheet(title=sheet_name)
        with zip_ref.open(file_name, 'r') as txt_file:
            lines = txt_file.read().decode('cp866').splitlines()
            for row_idx, line in enumerate(lines, start=1):
                values = line.strip().split(delimiter)
                for col_idx, cell_value in enumerate(values, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    process_folder(input_directory)

    if 'Sheet' in wb_txt.sheetnames:
        wb_txt.remove(wb_txt['Sheet'])
    if 'Sheet' in wb_other.sheetnames:
        wb_other.remove(wb_other['Sheet'])

    wb_txt.save(filename=output_txt)
    wb_other.save(filename=output_other)

def main():
    root = Tk()
    root.withdraw()
    
    input_directory = filedialog.askdirectory(title="Выберите директорию с файлами")
    if not input_directory:
        return
    
    output_txt = "txt_output.xlsx"
    output_other = "other_output.xlsx"
    delimiter = simpledialog.askstring("Ввод разделителя", "Введите разделитель:")
    if delimiter is None:
        return
    
    text_to_excel(input_directory, output_txt, output_other, delimiter)

if __name__ == "__main__":
    main()